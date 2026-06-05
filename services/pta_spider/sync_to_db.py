"""
PTA 爬取结果 → 数据库同步脚本
解析 爬取结果/ 目录下的数据，结构化存储到 ptadatabase

同步策略：
- experiment: 按名称 upsert，新增插入，已有更新
- problems_sets: 按 experiment_id upsert 题目内容
- submit_situation: 从提交记录.csv 解析，PTA用户ID→学号映射
- student + score: 从 PAPER_TRANSCRIPT.xlsx 解析
- student_code: 从 ANSWER_SHEET.zip 解析 (HTML→代码+测试结果)
- AI 相关表: 不动

数据源：
- 爬取结果/{实验名}/题目内容.txt → problems_sets
- 爬取结果/{实验名}/提交记录.csv → submit_situation
- 爬取结果/{实验名}/导出/*PAPER_TRANSCRIPT*.xlsx → score + student
- 爬取结果/{实验名}/导出/*SCORED_CODE*.zip → PTA用户ID映射
- 爬取结果/{实验名}/导出/*ANSWER_SHEET*.zip → student_code (代码+测试点结果)
"""
import os
import sys
import csv
import io
import posixpath
import struct
import zlib
import zipfile
import re
import traceback
from pathlib import Path
from datetime import datetime
from html import unescape
from html.parser import HTMLParser
import xml.etree.ElementTree as ET

import pymysql
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
DEFAULT_RUNTIME_PARENT = (
    Path(os.getenv("LOCALAPPDATA", str(Path.home())))
    if os.name == "nt"
    else Path.home() / ".local" / "share"
)
RUNTIME_DIR = Path(
    os.getenv("PTA_RUNTIME_DIR", str(DEFAULT_RUNTIME_PARENT / "teaching-assistant-platform" / "pta-spider"))
).resolve()

# Windows 终端 UTF-8 输出修复
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

# 加载 .env
_env_candidates = [
    RUNTIME_DIR / ".env",
    Path(".env"),
]
for _env in _env_candidates:
    if _env.exists():
        load_dotenv(_env)
        break
else:
    load_dotenv()

CRAWL_DIR = Path(os.getenv("PTA_CRAWL_DIR", str(PROJECT_ROOT / "爬取结果"))).resolve()


# ==================== ZIP 编码兼容工具 ====================

def _safe_zip_read(zf: zipfile.ZipFile, info: zipfile.ZipInfo) -> bytes:
    """
    安全读取 zip 内文件，绕过 Python 3.10 的文件名一致性校验。
    PTA 导出的 zip 中央目录用 GBK 编码，本地文件头用 UTF-8，
    导致 zipfile.read() 抛出 BadZipFile。
    此函数在校验失败时直接从底层读取压缩数据并解压。
    """
    try:
        return zf.read(info)
    except zipfile.BadZipFile as e:
        if "differ" not in str(e):
            raise
    # 底层读取: seek 到 local file header，跳过头部，读压缩数据
    zf.fp.seek(info.header_offset)
    fheader = zf.fp.read(30)
    if len(fheader) < 30 or fheader[:4] != b'PK\x03\x04':
        raise zipfile.BadZipFile("Bad local file header")
    fname_len, extra_len = struct.unpack('<HH', fheader[26:30])
    zf.fp.read(fname_len + extra_len)  # 跳过文件名和 extra
    comp_data = zf.fp.read(info.compress_size)
    if info.compress_type == zipfile.ZIP_STORED:
        return comp_data
    elif info.compress_type == zipfile.ZIP_DEFLATED:
        return zlib.decompress(comp_data, -15)
    else:
        raise zipfile.BadZipFile(f"Unsupported compression: {info.compress_type}")


def _decode_zip_filename(raw_name: str) -> str:
    """将 cp437 编码的 zip 文件名解码为 GBK (中文)"""
    try:
        return raw_name.encode('cp437').decode('gbk')
    except (UnicodeDecodeError, UnicodeEncodeError):
        return raw_name


_XLSX_NS_MAIN = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_PKG_REL_NS = {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}


def _cell_ref_to_index(cell_ref: str) -> int:
    col = 0
    for ch in cell_ref:
        if "A" <= ch <= "Z":
            col = col * 26 + (ord(ch) - ord("A") + 1)
        elif "a" <= ch <= "z":
            col = col * 26 + (ord(ch) - ord("a") + 1)
        else:
            break
    return max(col - 1, 0)


def _xlsx_get_text(node) -> str:
    if node is None:
        return ""
    parts = []
    for text_node in node.findall(".//x:t", _XLSX_NS_MAIN):
        parts.append(text_node.text or "")
    return "".join(parts) if parts else "".join(node.itertext())


def _read_xlsx_rows(xlsx_path: Path):
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        shared_strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            shared_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            shared_strings = [_xlsx_get_text(item) for item in shared_root.findall("x:si", _XLSX_NS_MAIN)]

        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {
            rel.attrib.get("Id"): rel.attrib.get("Target", "")
            for rel in rels_root.findall("pr:Relationship", _PKG_REL_NS)
        }

        first_sheet = workbook_root.find("x:sheets/x:sheet", _XLSX_NS_MAIN)
        if first_sheet is None:
            return []
        rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rel_map.get(rel_id, "")
        if not target:
            return []
        sheet_path = target if target.startswith("xl/") else posixpath.normpath(posixpath.join("xl", target))

        sheet_root = ET.fromstring(zf.read(sheet_path))
        rows = []
        for row in sheet_root.findall("x:sheetData/x:row", _XLSX_NS_MAIN):
            values = []
            for cell in row.findall("x:c", _XLSX_NS_MAIN):
                col_idx = _cell_ref_to_index(cell.attrib.get("r", "A1"))
                while len(values) <= col_idx:
                    values.append(None)

                cell_type = cell.attrib.get("t", "")
                if cell_type == "inlineStr":
                    value = _xlsx_get_text(cell.find("x:is", _XLSX_NS_MAIN))
                else:
                    raw = cell.findtext("x:v", default="", namespaces=_XLSX_NS_MAIN)
                    if cell_type == "s":
                        try:
                            value = shared_strings[int(raw)]
                        except Exception:
                            value = raw
                    else:
                        value = raw

                if isinstance(value, str):
                    value = value.strip()
                values[col_idx] = value if value != "" else None
            rows.append(tuple(values))
        return rows


class _AnswerSheetParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=False)
        self.pre_blocks = []
        self._pre_buffer = []
        self._capture_pre = False
        self.headers = []
        self.rows = []
        self._inside_result_table = False
        self._current_row = None
        self._current_cell = None
        self._current_cell_tag = None

    def handle_starttag(self, tag, attrs):
        attr_map = dict(attrs)
        class_tokens = set((attr_map.get("class") or "").split())

        if tag == "pre" and {"font-mono", "whitespace-pre-wrap"}.issubset(class_tokens):
            self._capture_pre = True
            self._pre_buffer = []
            return

        if tag == "table" and "result" in class_tokens:
            self._inside_result_table = True
            return

        if not self._inside_result_table:
            return

        if tag == "tr":
            self._current_row = []
        elif tag in ("th", "td"):
            self._current_cell = []
            self._current_cell_tag = tag

    def handle_endtag(self, tag):
        if tag == "pre" and self._capture_pre:
            text = unescape("".join(self._pre_buffer)).strip()
            self.pre_blocks.append(text)
            self._capture_pre = False
            self._pre_buffer = []
            return

        if not self._inside_result_table:
            return

        if tag in ("th", "td") and self._current_cell is not None:
            text = unescape("".join(self._current_cell)).strip()
            if self._current_row is not None:
                self._current_row.append(text)
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None:
            if self._current_row:
                if self.headers and self._current_cell_tag != "th":
                    self.rows.append(self._current_row)
                elif not self.headers:
                    self.headers = self._current_row
                else:
                    self.rows.append(self._current_row)
            self._current_row = None
            self._current_cell_tag = None
        elif tag == "table":
            self._inside_result_table = False

    def handle_data(self, data):
        if self._capture_pre:
            self._pre_buffer.append(data)
        elif self._current_cell is not None:
            self._current_cell.append(data)


def _parse_answer_sheet_html(html_text: str):
    parser = _AnswerSheetParser()
    parser.feed(html_text)
    md_table = ""
    if parser.headers:
        md_table = "| " + " | ".join(parser.headers) + " |\n"
        md_table += "| " + " | ".join(["---"] * len(parser.headers)) + " |\n"
        for row in parser.rows:
            md_table += "| " + " | ".join(row) + " |\n"
    return parser.pre_blocks, md_table


def get_db():
    return pymysql.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", "3306")),
        user=os.getenv("DB_USER", "root"),
        password=os.getenv("DB_PASSWORD", ""),
        database=os.getenv("DB_NAME", "ptadatabase"),
        charset="utf8mb4",
    )


# ==================== 1. 同步 experiment 表 ====================

def sync_experiments(conn):
    """扫描爬取结果目录，为每个实验文件夹创建/更新 experiment 记录"""
    cursor = conn.cursor()
    if not CRAWL_DIR.exists():
        print(f"[跳过] 爬取结果目录不存在: {CRAWL_DIR}")
        return {}

    exp_dirs = sorted([d for d in CRAWL_DIR.iterdir() if d.is_dir()])
    exp_map = {}  # name → experiment_id

    for idx, d in enumerate(exp_dirs, start=1):
        name = d.name
        cursor.execute("SELECT experiment_id FROM experiment WHERE name = %s", (name,))
        row = cursor.fetchone()
        if row:
            exp_map[name] = row[0]
        else:
            topic_sum = 0
            problem_file = d / "题目内容.txt"
            if problem_file.exists():
                content = problem_file.read_text(encoding="utf-8")
                topic_sum = len(re.findall(r'^\[\d+\]', content, re.MULTILINE))

            cursor.execute(
                "INSERT INTO experiment (num, name, topic_sum) VALUES (%s, %s, %s)",
                (idx, name, topic_sum)
            )
            exp_map[name] = cursor.lastrowid
            print(f"  [新增] experiment: {name} (id={exp_map[name]}, 题目数={topic_sum})")

    conn.commit()
    print(f"[OK] experiment 同步完成: {len(exp_map)} 个实验")
    return exp_map


# ==================== 2. 同步 problems_sets 表 ====================

def sync_problems(conn, exp_map):
    """将题目内容.txt 同步到 problems_sets 表"""
    cursor = conn.cursor()
    count = 0
    for name, eid in exp_map.items():
        problem_file = CRAWL_DIR / name / "题目内容.txt"
        if not problem_file.exists():
            continue
        content = problem_file.read_text(encoding="utf-8")
        cursor.execute(
            """INSERT INTO problems_sets (experiment_id, experiment_name, problem)
               VALUES (%s, %s, %s)
               ON DUPLICATE KEY UPDATE experiment_name=VALUES(experiment_name), problem=VALUES(problem)""",
            (eid, name, content)
        )
        count += 1
    conn.commit()
    print(f"[OK] problems_sets 同步完成: {count} 个")


# ==================== 3. 从 SCORED_CODE.zip 构建 PTA用户ID→学号 映射 ====================

def build_pta_user_map(exp_map):
    """
    从 SCORED_CODE.zip 的文件名中提取 PTA用户ID → 学号 的映射。
    文件名格式: {problem_id}_{problem_name}/{student_id}_{pta_user_id}_{problem_id}.c
    注意: 只需要读文件名（infolist），不需要读文件内容，所以不受编码 bug 影响。
    """
    pta_to_student = {}  # pta_user_id → student_id

    for name in exp_map:
        export_dir = CRAWL_DIR / name / "导出"
        if not export_dir.exists():
            continue
        zip_files = list(export_dir.glob("*SCORED_CODE*.zip"))
        if not zip_files:
            continue
        try:
            with zipfile.ZipFile(zip_files[0], "r") as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    fname = _decode_zip_filename(info.filename)
                    basename = fname.split("/")[-1] if "/" in fname else fname
                    # 格式: {student_id}_{pta_user_id}_{problem_id}.ext
                    parts = basename.split("_")
                    if len(parts) >= 2:
                        sid = parts[0]
                        pta_uid = parts[1]
                        if sid.isdigit() and 7 <= len(sid) <= 15:
                            if pta_uid.isdigit() and len(pta_uid) > 15:
                                pta_to_student[pta_uid] = sid
        except Exception as e:
            print(f"  [警告] 解析 SCORED_CODE.zip 映射失败 ({name}): {e}")

    print(f"[OK] PTA用户ID映射: {len(pta_to_student)} 个用户")
    return pta_to_student


# ==================== 4. 从成绩单 xlsx 构建 学号→姓名 映射 ====================

def build_student_name_map(exp_map):
    """从 PAPER_TRANSCRIPT.xlsx 提取 学号→姓名 映射"""
    try:
        import openpyxl
    except ImportError:
        print("[跳过] openpyxl 未安装")
        return {}

    student_names = {}  # student_id → student_name

    for name in exp_map:
        export_dir = CRAWL_DIR / name / "导出"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue
        try:
            wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            # 找到表头行 (包含 "学号" 或 "姓名" 的行)
            header_idx = None
            sid_col = None
            name_col = None
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    if val and "学号" in str(val):
                        header_idx = i
                        sid_col = j
                    if val and "姓名" in str(val):
                        header_idx = i
                        name_col = j
                if header_idx is not None:
                    break

            if header_idx is None:
                # 默认: PTA成绩单格式 Row2 是表头, Col1=学号, Col2=姓名
                header_idx = 2
                sid_col = 1
                name_col = 2

            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                sname = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
                if sid and sid != "None":
                    student_names[sid] = sname
            wb.close()
        except Exception as e:
            print(f"  [警告] 解析成绩单映射失败 ({name}): {e}")

    print(f"[OK] 学号→姓名映射: {len(student_names)} 个学生")
    return student_names


# ==================== 5. 同步提交记录 ====================

def sync_submissions(conn, exp_map, pta_to_student, student_names):
    """
    从提交记录.csv 同步到 submit_situation 表。
    CSV 中的用户ID是PTA内部ID，需要通过映射转换为学号。
    """
    cursor = conn.cursor()
    total = 0
    unmapped = set()

    for name, eid in exp_map.items():
        csv_file = CRAWL_DIR / name / "提交记录.csv"
        if not csv_file.exists():
            continue

        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            batch = []
            for row in reader:
                pta_uid = row.get("用户ID", "").strip()
                problem_id = row.get("题目ID", "").strip()
                status = row.get("状态", "")
                score = row.get("分数", "0")
                compiler = row.get("编译器", "")
                time_used = row.get("用时", "0")
                memory = row.get("内存", "0")
                submit_time = row.get("提交时间", "")

                # PTA用户ID → 学号
                student_id = pta_to_student.get(pta_uid, pta_uid)
                if student_id == pta_uid and len(pta_uid) > 15:
                    unmapped.add(pta_uid)

                student_name = student_names.get(student_id, "")

                # 内存: PTA返回字节，转KB
                try:
                    memory_kb = str(int(float(memory)) // 1024) if memory else "0"
                except (ValueError, TypeError):
                    memory_kb = "0"

                # 用时: PTA返回秒，转ms
                try:
                    runtime_ms = str(int(float(time_used) * 1000)) if time_used else "0"
                except (ValueError, TypeError):
                    runtime_ms = "0"

                batch.append((
                    submit_time, status, score, problem_id,
                    eid, name, runtime_ms, memory_kb,
                    student_id, student_name
                ))

            if batch:
                cursor.executemany(
                    """INSERT INTO submit_situation
                       (submit_time, situation, score, serial_number,
                        experiment_id, experiment_name, runtime_ms, memory_kb,
                        student_id, student_name)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                         situation=VALUES(situation),
                         score=VALUES(score),
                         experiment_name=VALUES(experiment_name),
                         runtime_ms=VALUES(runtime_ms),
                         memory_kb=VALUES(memory_kb),
                         student_name=VALUES(student_name)""",
                    batch
                )
                total += len(batch)

    conn.commit()
    if unmapped:
        print(f"  [注意] {len(unmapped)} 个PTA用户ID未找到学号映射")
    print(f"[OK] submit_situation 同步完成: {total} 条")


# ==================== 6. 从成绩单 xlsx 同步 score + student ====================

def sync_transcript(conn, exp_map):
    """
    从 PAPER_TRANSCRIPT.xlsx 解析成绩单，同步到 score 表和 student 表。
    PTA成绩单格式:
      Row 0: 标题行 (成绩明细 - xxx - 全体考生)
      Row 1: 题目标号行
      Row 2: 表头行 (用户组, 学号/邮箱、电话, 姓名/昵称, MOOCID, 总分, 排名, ...)
      Row 3+: 数据行
    """
    try:
        import openpyxl
    except ImportError:
        print("[跳过] openpyxl 未安装，无法解析成绩单")
        return

    cursor = conn.cursor()
    score_count = 0
    student_set = set()

    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "导出"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue

        try:
            wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 4:
                wb.close()
                continue

            # PTA 固定格式: Row2 是表头, Col1=学号, Col2=姓名, Col4=总分
            # 但为了健壮性，动态查找
            header_idx = 2  # 默认
            sid_col, name_col, score_col = 1, 2, 4  # 默认

            for i, row in enumerate(rows[:5]):
                for j, val in enumerate(row):
                    s = str(val) if val else ""
                    if "学号" in s:
                        header_idx = i
                        sid_col = j
                    if "姓名" in s:
                        name_col = j
                    if "总分" in s:
                        score_col = j

            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None":
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
                score_val = row[score_col] if score_col < len(row) else None

                # 同步 student 表
                if sid not in student_set:
                    cursor.execute(
                        "INSERT IGNORE INTO student (student_id, name) VALUES (%s, %s)",
                        (sid, sname)
                    )
                    student_set.add(sid)

                # 同步 score 表
                if score_val is not None:
                    try:
                        score_num = float(score_val)
                    except (ValueError, TypeError):
                        score_num = 0
                    cursor.execute(
                        """INSERT INTO score (username, real_name, experiment_id, score, status)
                           VALUES (%s, %s, %s, %s, 'graded')
                           ON DUPLICATE KEY UPDATE
                             real_name=VALUES(real_name),
                             score=VALUES(score),
                             status=VALUES(status)""",
                        (sid, sname, eid, score_num)
                    )
                    score_count += 1

            wb.close()
        except Exception as e:
            print(f"  [警告] 解析成绩单失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] score 同步完成: {score_count} 条, student 新增: {len(student_set)} 人")


# ==================== 7. 从 ANSWER_SHEET.zip 同步 student_code ====================

def sync_student_code(conn, exp_map):
    """
    从 ANSWER_SHEET.zip 解析学生代码+测试结果，同步到 student_code 表。
    zip 内结构: {班级名}/html/{学号}-{姓名}.html
    每个 HTML 包含多道题的代码和测试点结果表格。
    使用 BeautifulSoup 提取:
      - 代码: pre.box.text-xs.font-mono.whitespace-pre-wrap
      - 测试结果: table.result → Markdown 表格
    输出格式与老项目一致: "第N题如下:\n{代码}\n{测试结果表格}\n"
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("[跳过] beautifulsoup4 未安装，无法解析答题情况")
        return

    cursor = conn.cursor()
    count = 0

    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "导出"
        if not export_dir.exists():
            continue
        zip_files = list(export_dir.glob("*ANSWER_SHEET*.zip"))
        if not zip_files:
            continue

        try:
            with zipfile.ZipFile(zip_files[0], "r") as zf:
                html_infos = [i for i in zf.infolist()
                              if not i.is_dir() and i.filename.endswith('.html')]

                for info in html_infos:
                    decoded_name = _decode_zip_filename(info.filename)
                    # 文件名格式: {班级名}/html/{学号}-{姓名}.html
                    basename = decoded_name.split("/")[-1] if "/" in decoded_name else decoded_name
                    stem = basename.replace(".html", "")
                    parts = stem.split("-", 1)
                    if len(parts) < 2:
                        continue
                    sid = parts[0].strip()
                    sname = parts[1].strip()
                    if not sid.isdigit() or len(sid) < 7:
                        continue

                    # 读取 HTML
                    try:
                        raw = _safe_zip_read(zf, info)
                        html = raw.decode("utf-8", errors="replace")
                    except Exception as e:
                        print(f"    [警告] 读取 {decoded_name} 失败: {e}")
                        continue

                    # 解析 HTML
                    soup = BeautifulSoup(html, "html.parser")
                    pre_elements = soup.select("pre.box.text-xs.font-mono.whitespace-pre-wrap")

                    # 提取测试结果表格 → Markdown
                    md_table = ""
                    table = soup.find("table", {"class": "result"})
                    if table:
                        thead = table.find("thead")
                        if thead:
                            headers = [th.get_text(strip=True) for th in thead.find_all("th")]
                            md_table = "| " + " | ".join(headers) + " |\n"
                            md_table += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                        tbody = table.find("tbody")
                        if tbody:
                            for tr in tbody.find_all("tr"):
                                cells = [td.get_text(strip=True) for td in tr.find_all("td")]
                                md_table += "| " + " | ".join(cells) + " |\n"

                    # 合并代码: 老项目格式 — 奇数 pre 是代码，偶数 pre 跳过
                    merged = ""
                    q_num = 1
                    for idx, pre in enumerate(pre_elements):
                        if idx % 2 == 0:  # 奇数位(0-indexed偶数)是代码
                            code_text = pre.get_text()
                            merged += f"第{q_num}题如下:\n{code_text}\n"
                            if md_table:
                                merged += md_table + "\n"
                            q_num += 1

                    if not merged.strip():
                        continue

                    cursor.execute(
                        """INSERT INTO student_code
                           (experiment_id, experiment_name, student_id, student_name, code)
                           VALUES (%s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE
                             experiment_name=VALUES(experiment_name),
                             student_name=VALUES(student_name),
                             code=VALUES(code)""",
                        (eid, name, sid, sname, merged)
                    )
                    count += 1

        except Exception as e:
            print(f"  [警告] 解析答题情况包失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] student_code 同步完成: {count} 条")


# ==================== 8. 从成绩单 xlsx 同步每题得分明细 ====================

def sync_problem_scores(conn, exp_map):
    """
    从 PAPER_TRANSCRIPT.xlsx 解析每题得分明细，同步到 problem_score_detail 表。
    用于计算正答率、题目难度分析、分数分布等。

    PTA成绩单格式:
      Row 0: 标题行 + 题目类型分组（单选题、编程题 等，合并单元格）
      Row 1: 题目标号行（2-1, 2-2, ..., 7-1, 7-2 等）
      Row 2: 表头行（用户组, 学号, 姓名, MOOCID, 总分(xx), 排名, 耗时, 答案/满分...）
      Row 3+: 数据行
    """
    try:
        import openpyxl
    except ImportError:
        print("[跳过] openpyxl 未安装，无法解析每题得分")
        return

    cursor = conn.cursor()

    # 先确保表存在
    try:
        cursor.execute("SELECT 1 FROM problem_score_detail LIMIT 1")
    except Exception:
        print("[跳过] problem_score_detail 表不存在，请先执行 V10 迁移")
        return

    count = 0
    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "导出"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue

        try:
            wb = openpyxl.load_workbook(xlsx_files[0], read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 4:
                wb.close()
                continue

            # Row 0: 题目类型分组（合并单元格，需要前向填充）
            row0 = list(rows[0])
            # Row 1: 题目标号
            row1 = list(rows[1])
            # Row 2: 表头（含满分信息）
            row2 = list(rows[2])

            # 找到学号列和姓名列
            sid_col, name_col, total_col, rank_col = 1, 2, 4, 5
            for j, val in enumerate(row2):
                s = str(val) if val else ""
                if "学号" in s:
                    sid_col = j
                if "姓名" in s:
                    name_col = j
                if "总分" in s:
                    total_col = j
                if "排名" in s:
                    rank_col = j

            # 构建题目列映射: col_index → {label, type, max_score}
            # 题目列从 row2 的表头后面开始（跳过 用户组/学号/姓名/MOOCID/总分/排名/耗时）
            problem_cols = {}
            # 前向填充 row0 的题目类型
            current_type = ""
            for j in range(len(row0)):
                if row0[j] is not None and str(row0[j]).strip():
                    current_type = str(row0[j]).strip()
                row0[j] = current_type

            for j in range(len(row1)):
                label = str(row1[j]).strip() if row1[j] is not None else ""
                if not label or label == "None" or label == "题目标号":
                    continue
                # 跳过汇总列（如 "单选题得分", "编程题得分"）
                if "得分" in label:
                    continue
                # 解析满分: row2[j] 格式如 "C(3.0)" 或 "10.0"
                header_val = str(row2[j]).strip() if j < len(row2) and row2[j] is not None else ""
                max_score = 0
                # 尝试从括号中提取满分: "C(3.0)" → 3.0
                import re as _re
                m = _re.search(r'\((\d+\.?\d*)\)', header_val)
                if m:
                    max_score = float(m.group(1))
                else:
                    # 直接是数字: "10.0"
                    try:
                        max_score = float(header_val)
                    except (ValueError, TypeError):
                        pass

                ptype = row0[j] if j < len(row0) else ""
                problem_cols[j] = {
                    "label": label,
                    "type": ptype,
                    "max_score": max_score,
                }

            if not problem_cols:
                wb.close()
                continue

            # 解析数据行
            batch = []
            for row in rows[3:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None" or not sid[0].isdigit():
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""

                # 总分和排名
                total_score = 0
                ranking = 0
                try:
                    total_score = float(row[total_col]) if total_col < len(row) and row[total_col] is not None else 0
                except (ValueError, TypeError):
                    pass
                try:
                    ranking = int(float(row[rank_col])) if rank_col < len(row) and row[rank_col] is not None else 0
                except (ValueError, TypeError):
                    pass

                for col_idx, pinfo in problem_cols.items():
                    if col_idx >= len(row):
                        continue
                    cell = row[col_idx]
                    actual = 0
                    if cell is not None and str(cell).strip() not in ("", "-", "None"):
                        try:
                            actual = float(cell)
                        except (ValueError, TypeError):
                            pass

                    batch.append((
                        eid, sid, sname,
                        pinfo["label"], pinfo["type"], pinfo["max_score"],
                        actual, total_score, ranking
                    ))

            if batch:
                cursor.executemany(
                    """INSERT INTO problem_score_detail
                       (experiment_id, student_id, student_name,
                        problem_label, problem_type, max_score,
                        actual_score, total_score, ranking)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                         student_name=VALUES(student_name),
                         problem_type=VALUES(problem_type),
                         max_score=VALUES(max_score),
                         actual_score=VALUES(actual_score),
                         total_score=VALUES(total_score),
                         ranking=VALUES(ranking)""",
                    batch
                )
                count += len(batch)

            wb.close()
        except Exception as e:
            print(f"  [警告] 解析每题得分失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] problem_score_detail 同步完成: {count} 条")


# ==================== 9. 清理检查（不删除 AI 数据） ====================

def get_db():
    config = {
        "host": os.getenv("DB_HOST", "localhost"),
        "port": int(os.getenv("DB_PORT", "3306")),
        "user": os.getenv("DB_USER", "root"),
        "password": os.getenv("DB_PASSWORD", ""),
        "database": os.getenv("DB_NAME", "ptadatabase"),
        "charset": "utf8mb4",
    }

    try:
        return pymysql.connect(**config)
    except RuntimeError as e:
        if "cryptography" not in str(e).lower():
            raise
        return pymysql.connect(**config, ssl={"ssl": True})


def _safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        text = str(value).strip()
        if not text or text in {"-", "None"}:
            return default
        return float(text)
    except (ValueError, TypeError):
        return default


def build_student_name_map(exp_map):
    student_names = {}

    for name in exp_map:
        export_dir = CRAWL_DIR / name / "瀵煎嚭"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue

        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            header_idx = None
            sid_col = None
            name_col = None
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    text = str(val) if val is not None else ""
                    if "瀛﹀彿" in text:
                        header_idx = i
                        sid_col = j
                    if "濮撳悕" in text:
                        header_idx = i
                        name_col = j
                if header_idx is not None:
                    break

            if header_idx is None:
                header_idx, sid_col, name_col = 2, 1, 2

            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None":
                    continue
                sname = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
                student_names[sid] = sname
        except Exception as e:
            print(f"  [警告] 解析成绩单姓名映射失败 ({name}): {e}")

    print(f"[OK] 学号-姓名映射: {len(student_names)} 个学生")
    return student_names


def sync_transcript(conn, exp_map):
    cursor = conn.cursor()
    score_count = 0
    student_set = set()

    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "瀵煎嚭"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue

        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            if len(rows) < 4:
                continue

            header_idx = 2
            sid_col, name_col, score_col = 1, 2, 4
            for i, row in enumerate(rows[:5]):
                for j, val in enumerate(row):
                    text = str(val) if val is not None else ""
                    if "瀛﹀彿" in text:
                        header_idx = i
                        sid_col = j
                    if "濮撳悕" in text:
                        name_col = j
                    if "鎬诲垎" in text:
                        score_col = j

            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None":
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
                score_num = _safe_float(row[score_col] if score_col < len(row) else None, 0)

                if sid not in student_set:
                    cursor.execute(
                        "INSERT IGNORE INTO student (student_id, name) VALUES (%s, %s)",
                        (sid, sname)
                    )
                    student_set.add(sid)

                cursor.execute(
                    """INSERT INTO score (username, real_name, experiment_id, score, status)
                       VALUES (%s, %s, %s, %s, 'graded')
                       ON DUPLICATE KEY UPDATE
                         real_name=VALUES(real_name),
                         score=VALUES(score),
                         status=VALUES(status)""",
                    (sid, sname, eid, score_num)
                )
                score_count += 1
        except Exception as e:
            print(f"  [警告] 解析成绩单失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] score 同步完成: {score_count} 条, student 新增: {len(student_set)} 人")


def sync_student_code(conn, exp_map):
    cursor = conn.cursor()
    count = 0

    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "瀵煎嚭"
        if not export_dir.exists():
            continue
        zip_files = list(export_dir.glob("*ANSWER_SHEET*.zip"))
        if not zip_files:
            continue

        try:
            with zipfile.ZipFile(zip_files[0], "r") as zf:
                html_infos = [i for i in zf.infolist() if not i.is_dir() and i.filename.endswith(".html")]
                for info in html_infos:
                    decoded_name = _decode_zip_filename(info.filename)
                    basename = decoded_name.split("/")[-1] if "/" in decoded_name else decoded_name
                    stem = basename.replace(".html", "")
                    parts = stem.split("-", 1)
                    if len(parts) < 2:
                        continue
                    sid = parts[0].strip()
                    sname = parts[1].strip()
                    if not sid.isdigit() or len(sid) < 7:
                        continue

                    try:
                        raw = _safe_zip_read(zf, info)
                        html_text = raw.decode("utf-8", errors="replace")
                    except Exception as e:
                        print(f"    [警告] 读取 {decoded_name} 失败: {e}")
                        continue

                    pre_blocks, md_table = _parse_answer_sheet_html(html_text)
                    merged = ""
                    q_num = 1
                    for idx, code_text in enumerate(pre_blocks):
                        if idx % 2 == 0:
                            merged += f"第{q_num}题如下:\n{code_text}\n"
                            if md_table:
                                merged += md_table + "\n"
                            q_num += 1

                    if not merged.strip():
                        continue

                    cursor.execute(
                        """INSERT INTO student_code
                           (experiment_id, experiment_name, student_id, student_name, code)
                           VALUES (%s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE
                             experiment_name=VALUES(experiment_name),
                             student_name=VALUES(student_name),
                             code=VALUES(code)""",
                        (eid, name, sid, sname, merged)
                    )
                    count += 1
        except Exception as e:
            print(f"  [警告] 解析答题情况包失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] student_code 同步完成: {count} 条")


def sync_problem_scores(conn, exp_map):
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT 1 FROM problem_score_detail LIMIT 1")
    except Exception:
        print("[跳过] problem_score_detail 表不存在，请先执行 V10 迁移")
        return

    count = 0
    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / "瀵煎嚭"
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue

        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            if len(rows) < 4:
                continue

            row0 = list(rows[0])
            row1 = list(rows[1])
            row2 = list(rows[2])

            sid_col, name_col, total_col, rank_col = 1, 2, 4, 5
            for j, val in enumerate(row2):
                text = str(val) if val is not None else ""
                if "瀛﹀彿" in text:
                    sid_col = j
                if "濮撳悕" in text:
                    name_col = j
                if "鎬诲垎" in text:
                    total_col = j
                if "鎺掑悕" in text:
                    rank_col = j

            current_type = ""
            for j in range(len(row0)):
                if row0[j] is not None and str(row0[j]).strip():
                    current_type = str(row0[j]).strip()
                row0[j] = current_type

            problem_cols = {}
            for j in range(len(row1)):
                label = str(row1[j]).strip() if row1[j] is not None else ""
                if not label or label == "None" or label == "棰樼洰鏍囧彿" or "寰楀垎" in label:
                    continue

                header_val = str(row2[j]).strip() if j < len(row2) and row2[j] is not None else ""
                m = re.search(r"\((\d+\.?\d*)\)", header_val)
                max_score = float(m.group(1)) if m else _safe_float(header_val, 0)
                problem_cols[j] = {
                    "label": label,
                    "type": row0[j] if j < len(row0) else "",
                    "max_score": max_score,
                }

            if not problem_cols:
                continue

            batch = []
            for row in rows[3:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None" or not sid[0].isdigit():
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
                total_score = _safe_float(row[total_col] if total_col < len(row) else None, 0)
                ranking = int(_safe_float(row[rank_col] if rank_col < len(row) else None, 0))

                for col_idx, pinfo in problem_cols.items():
                    if col_idx >= len(row):
                        continue
                    actual = _safe_float(row[col_idx], 0)
                    batch.append((
                        eid, sid, sname,
                        pinfo["label"], pinfo["type"], pinfo["max_score"],
                        actual, total_score, ranking
                    ))

            if batch:
                cursor.executemany(
                    """INSERT INTO problem_score_detail
                       (experiment_id, student_id, student_name,
                        problem_label, problem_type, max_score,
                        actual_score, total_score, ranking)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                         student_name=VALUES(student_name),
                         problem_type=VALUES(problem_type),
                         max_score=VALUES(max_score),
                         actual_score=VALUES(actual_score),
                         total_score=VALUES(total_score),
                         ranking=VALUES(ranking)""",
                    batch
                )
                count += len(batch)
        except Exception as e:
            print(f"  [警告] 解析每题得分失败 ({name}): {e}")

    conn.commit()
    print(f"[OK] problem_score_detail 同步完成: {count} 条")


def cleanup_check(conn, exp_map):
    """检查数据库中有但爬取结果中没有的实验，仅提示不删除"""
    cursor = conn.cursor()
    cursor.execute("SELECT experiment_id, name FROM experiment")
    db_experiments = {row[1]: row[0] for row in cursor.fetchall()}

    crawl_names = set(exp_map.keys())
    orphan_names = set(db_experiments.keys()) - crawl_names

    if not orphan_names:
        print("[OK] 没有过期实验")
        return

    print(f"[注意] {len(orphan_names)} 个实验在数据库中但不在爬取结果中:")
    for n in sorted(orphan_names):
        print(f"  - {n} (id={db_experiments[n]})")
    print("  (仅提示，不自动删除。AI数据始终保留)")


# ==================== 9. 统计报告 ====================

def print_stats(conn):
    cursor = conn.cursor()
    tables = [
        "experiment", "student", "problems_sets",
        "submit_situation", "score", "student_code",
        "problem_score_detail",
        "ai_remarks", "ai_suggested_problems", "ai_remark_organization"
    ]
    print(f"\n{'='*40}")
    print("数据库统计:")
    print(f"{'='*40}")
    for t in tables:
        try:
            cursor.execute(f"SELECT COUNT(*) FROM `{t}`")
            cnt = cursor.fetchone()[0]
            print(f"  {t:30s} {cnt:>8} 条")
        except Exception:
            print(f"  {t:30s} (表不存在)")
    print(f"{'='*40}")


# ==================== 主入口 ====================

def sync_all(crawl_dir=None):
    """执行全量同步"""
    global CRAWL_DIR
    if crawl_dir:
        CRAWL_DIR = Path(crawl_dir)

    print(f"\n{'='*50}")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] PTA 数据同步开始")
    print(f"爬取结果目录: {CRAWL_DIR}")
    print(f"{'='*50}\n")

    conn = get_db()
    try:
        # 1. 同步实验
        exp_map = sync_experiments(conn)
        if not exp_map:
            print("[结束] 没有找到任何实验数据")
            return

        # 2. 同步题目内容
        sync_problems(conn, exp_map)

        # 3. 构建 PTA用户ID→学号 映射 (从 SCORED_CODE.zip)
        pta_to_student = build_pta_user_map(exp_map)

        # 4. 构建 学号→姓名 映射 (从 PAPER_TRANSCRIPT.xlsx)
        student_names = build_student_name_map(exp_map)

        # 5. 同步成绩单 + 学生信息 (先于提交记录，确保 student 表有数据)
        sync_transcript(conn, exp_map)

        # 6. 同步提交记录 (使用映射转换 PTA用户ID→学号)
        sync_submissions(conn, exp_map, pta_to_student, student_names)

        # 7. 同步学生代码
        sync_student_code(conn, exp_map)

        # 8. 同步每题得分明细（正答率数据源）
        sync_problem_scores(conn, exp_map)

        # 9. 检查过期数据
        cleanup_check(conn, exp_map)

        # 9. 统计
        print_stats(conn)

        print(f"\n[完成] 数据同步成功")

    except Exception as e:
        print(f"\n[错误] 同步失败: {e}")
        traceback.print_exc()
    finally:
        conn.close()


_PTA_EXPORT_DIR = "\u5bfc\u51fa"
_HEADER_STUDENT_NO = "\u5b66\u53f7"
_HEADER_STUDENT_NAME = "\u59d3\u540d"
_HEADER_TOTAL_SCORE = "\u603b\u5206"
_HEADER_RANK = "\u6392\u540d"
_HEADER_PROBLEM_LABEL = "\u9898\u76ee\u6807\u53f7"
_HEADER_SCORE = "\u5f97\u5206"


def _normalize_experiment_names(experiment_names):
    if not experiment_names:
        return None
    return {str(name).strip() for name in experiment_names if str(name).strip()}


def build_student_name_map(exp_map):
    student_names = {}
    for name in exp_map:
        export_dir = CRAWL_DIR / name / _PTA_EXPORT_DIR
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue
        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            header_idx = None
            sid_col = None
            name_col = None
            for i, row in enumerate(rows):
                for j, val in enumerate(row):
                    text = str(val) if val is not None else ""
                    if _HEADER_STUDENT_NO in text:
                        header_idx = i
                        sid_col = j
                    if _HEADER_STUDENT_NAME in text:
                        header_idx = i
                        name_col = j
                if header_idx is not None:
                    break
            if header_idx is None:
                header_idx, sid_col, name_col = 2, 1, 2
            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None":
                    continue
                sname = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""
                student_names[sid] = sname
        except Exception as e:
            print(f"  [警告] 解析成绩单姓名映射失败 ({name}): {e}")
    print(f"[OK] 学号-姓名映射: {len(student_names)} 个学生")
    return student_names


def sync_transcript(conn, exp_map):
    cursor = conn.cursor()
    score_count = 0
    student_set = set()
    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / _PTA_EXPORT_DIR
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue
        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            if len(rows) < 4:
                continue
            header_idx = 2
            sid_col, name_col, score_col = 1, 2, 4
            for i, row in enumerate(rows[:5]):
                for j, val in enumerate(row):
                    text = str(val) if val is not None else ""
                    if _HEADER_STUDENT_NO in text:
                        header_idx = i
                        sid_col = j
                    if _HEADER_STUDENT_NAME in text:
                        name_col = j
                    if _HEADER_TOTAL_SCORE in text:
                        score_col = j
            for row in rows[header_idx + 1:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None":
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
                score_num = _safe_float(row[score_col] if score_col < len(row) else None, 0)
                if sid not in student_set:
                    cursor.execute("INSERT IGNORE INTO student (student_id, name) VALUES (%s, %s)", (sid, sname))
                    student_set.add(sid)
                cursor.execute(
                    """INSERT INTO score (username, real_name, experiment_id, score, status)
                       VALUES (%s, %s, %s, %s, 'graded')
                       ON DUPLICATE KEY UPDATE
                         real_name=VALUES(real_name),
                         score=VALUES(score),
                         status=VALUES(status)""",
                    (sid, sname, eid, score_num)
                )
                score_count += 1
        except Exception as e:
            print(f"  [警告] 解析成绩单失败 ({name}): {e}")
    conn.commit()
    print(f"[OK] score 同步完成: {score_count} 条, student 新增: {len(student_set)} 人")


def sync_student_code(conn, exp_map):
    cursor = conn.cursor()
    count = 0
    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / _PTA_EXPORT_DIR
        if not export_dir.exists():
            continue
        zip_files = list(export_dir.glob("*ANSWER_SHEET*.zip"))
        if not zip_files:
            continue
        try:
            with zipfile.ZipFile(zip_files[0], "r") as zf:
                html_infos = [i for i in zf.infolist() if not i.is_dir() and i.filename.endswith(".html")]
                for info in html_infos:
                    decoded_name = _decode_zip_filename(info.filename)
                    basename = decoded_name.split("/")[-1] if "/" in decoded_name else decoded_name
                    stem = basename[:-5] if basename.lower().endswith(".html") else basename
                    parts = stem.split("-", 1)
                    if len(parts) < 2:
                        continue
                    sid = parts[0].strip()
                    sname = parts[1].strip()
                    if not sid.isdigit() or len(sid) < 7:
                        continue
                    raw = _safe_zip_read(zf, info)
                    html_text = raw.decode("utf-8", errors="replace")
                    pre_blocks, md_table = _parse_answer_sheet_html(html_text)
                    merged = []
                    q_num = 1
                    for idx, code_text in enumerate(pre_blocks):
                        if idx % 2 == 0:
                            merged.append(f"\u7b2c{q_num}\u9898\u5982\u4e0b:\n{code_text}")
                            if md_table:
                                merged.append(md_table.rstrip())
                            q_num += 1
                    if not merged:
                        continue
                    cursor.execute(
                        """INSERT INTO student_code
                           (experiment_id, experiment_name, student_id, student_name, code)
                           VALUES (%s, %s, %s, %s, %s)
                           ON DUPLICATE KEY UPDATE
                             experiment_name=VALUES(experiment_name),
                             student_name=VALUES(student_name),
                             code=VALUES(code)""",
                        (eid, name, sid, sname, "\n".join(merged) + "\n")
                    )
                    count += 1
        except Exception as e:
            print(f"  [警告] 解析答题情况包失败 ({name}): {e}")
    conn.commit()
    print(f"[OK] student_code 同步完成: {count} 条")


def sync_problem_scores(conn, exp_map):
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT 1 FROM problem_score_detail LIMIT 1")
    except Exception:
        print("[跳过] problem_score_detail 表不存在，请先执行 V10 迁移")
        return
    count = 0
    for name, eid in exp_map.items():
        export_dir = CRAWL_DIR / name / _PTA_EXPORT_DIR
        if not export_dir.exists():
            continue
        xlsx_files = list(export_dir.glob("*PAPER_TRANSCRIPT*.xlsx"))
        if not xlsx_files:
            continue
        try:
            rows = _read_xlsx_rows(xlsx_files[0])
            if len(rows) < 4:
                continue
            row0 = list(rows[0])
            row1 = list(rows[1])
            row2 = list(rows[2])
            sid_col, name_col, total_col, rank_col = 1, 2, 4, 5
            for j, val in enumerate(row2):
                text = str(val) if val is not None else ""
                if _HEADER_STUDENT_NO in text:
                    sid_col = j
                if _HEADER_STUDENT_NAME in text:
                    name_col = j
                if _HEADER_TOTAL_SCORE in text:
                    total_col = j
                if _HEADER_RANK in text:
                    rank_col = j
            current_type = ""
            for j in range(len(row0)):
                if row0[j] is not None and str(row0[j]).strip():
                    current_type = str(row0[j]).strip()
                row0[j] = current_type
            problem_cols = {}
            for j in range(len(row1)):
                label = str(row1[j]).strip() if row1[j] is not None else ""
                if not label or label == "None" or label == _HEADER_PROBLEM_LABEL or _HEADER_SCORE in label:
                    continue
                header_val = str(row2[j]).strip() if j < len(row2) and row2[j] is not None else ""
                m = re.search(r"\((\d+\.?\d*)\)", header_val)
                max_score = float(m.group(1)) if m else _safe_float(header_val, 0)
                problem_cols[j] = {"label": label, "type": row0[j] if j < len(row0) else "", "max_score": max_score}
            if not problem_cols:
                continue
            batch = []
            for row in rows[3:]:
                if not row or sid_col >= len(row) or not row[sid_col]:
                    continue
                sid = str(row[sid_col]).strip()
                if not sid or sid == "None" or not sid[0].isdigit():
                    continue
                sname = str(row[name_col]).strip() if name_col < len(row) and row[name_col] else ""
                total_score = _safe_float(row[total_col] if total_col < len(row) else None, 0)
                ranking = int(_safe_float(row[rank_col] if rank_col < len(row) else None, 0))
                for col_idx, pinfo in problem_cols.items():
                    if col_idx >= len(row):
                        continue
                    batch.append((
                        eid, sid, sname,
                        pinfo["label"], pinfo["type"], pinfo["max_score"],
                        _safe_float(row[col_idx], 0), total_score, ranking
                    ))
            if batch:
                cursor.executemany(
                    """INSERT INTO problem_score_detail
                       (experiment_id, student_id, student_name,
                        problem_label, problem_type, max_score,
                        actual_score, total_score, ranking)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       ON DUPLICATE KEY UPDATE
                         student_name=VALUES(student_name),
                         problem_type=VALUES(problem_type),
                         max_score=VALUES(max_score),
                         actual_score=VALUES(actual_score),
                         total_score=VALUES(total_score),
                         ranking=VALUES(ranking)""",
                    batch
                )
                count += len(batch)
        except Exception as e:
            print(f"  [警告] 解析每题得分失败 ({name}): {e}")
    conn.commit()
    print(f"[OK] problem_score_detail 同步完成: {count} 条")


def sync_all(crawl_dir=None, strict=True, experiment_names=None):
    global CRAWL_DIR
    if crawl_dir:
        CRAWL_DIR = Path(crawl_dir)
    selected_names = _normalize_experiment_names(experiment_names)

    print(f"\n{'=' * 50}")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] PTA 数据同步开始")
    print(f"爬取结果目录: {CRAWL_DIR}")
    if selected_names:
        print(f"限定同步实验: {', '.join(sorted(selected_names))}")
    print(f"{'=' * 50}\n")

    conn = get_db()
    report = {
        "ok": False,
        "crawl_dir": str(CRAWL_DIR),
        "experiment_filter": sorted(selected_names) if selected_names else None,
        "experiment_count": 0,
    }

    try:
        exp_map = sync_experiments(conn)
        if selected_names:
            exp_map = {name: eid for name, eid in exp_map.items() if name in selected_names}
        report["experiment_count"] = len(exp_map)
        if not exp_map:
            message = f"No experiment data found in crawl directory: {CRAWL_DIR}"
            if selected_names:
                message = f"No selected experiment data found: {sorted(selected_names)}"
            print(f"[错误] {message}")
            report["error"] = message
            if strict:
                raise RuntimeError(message)
            return report

        sync_problems(conn, exp_map)
        pta_to_student = build_pta_user_map(exp_map)
        student_names = build_student_name_map(exp_map)
        sync_transcript(conn, exp_map)
        sync_submissions(conn, exp_map, pta_to_student, student_names)
        sync_student_code(conn, exp_map)
        sync_problem_scores(conn, exp_map)
        cleanup_check(conn, exp_map)
        print_stats(conn)

        print("\n[完成] 数据同步成功")
        report["ok"] = True
        return report
    except Exception as e:
        print(f"\n[错误] 同步失败: {e}")
        traceback.print_exc()
        report["error"] = str(e)
        if strict:
            raise
        return report
    finally:
        conn.close()


if __name__ == "__main__":
    crawl_dir = sys.argv[1] if len(sys.argv) > 1 else None
    sync_all(crawl_dir)
